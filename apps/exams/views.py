from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import BankSoal, ButirSoal, Ujian, SesiUjian, JawabanSiswa, JadwalPengawas
from .forms import BankSoalForm, ButirSoalForm, UjianForm, JadwalPengawasForm
from openpyxl import load_workbook
from django.http import HttpResponse, JsonResponse
import json
from django.utils import timezone
from django.db.models import Sum
import openpyxl
from io import BytesIO
from reportlab.lib.units import cm

@login_required
def bank_soal_list(request):
    if request.user.role == 'guru':
        bank_soal_list = BankSoal.objects.filter(guru=request.user)
    else:
        bank_soal_list = BankSoal.objects.all()
    return render(request, 'exams/bank_soal_list.html', {'bank_soal_list': bank_soal_list})

@login_required
def bank_soal_create(request):
    if request.user.role != 'guru':
        return redirect('login')
    
    if request.method == 'POST':
        form = BankSoalForm(request.POST, user=request.user)
        if form.is_valid():
            bank = form.save(commit=False)
            bank.guru = request.user
            bank.save()
            messages.success(request, 'Bank Soal berhasil dibuat. Silakan tambah butir soal.')
            return redirect('bank_soal_detail', pk=bank.pk)
    else:
        form = BankSoalForm(user=request.user)
    
    # Get all mapels that are NOT in user's ampu_mapel
    from apps.academic.models import MataPelajaran
    existing_ids = request.user.ampu_mapel.values_list('id', flat=True)
    available_mapels = MataPelajaran.objects.exclude(id__in=existing_ids).order_by('nama')
    
    return render(request, 'exams/bank_soal_form.html', {
        'form': form, 
        'title': 'Buat Bank Soal',
        'available_mapels': available_mapels
    })

@login_required
def bank_soal_edit(request, pk):
    bank = get_object_or_404(BankSoal, pk=pk)
    if request.user.role == 'guru' and bank.guru != request.user:
        return redirect('login')
        
    if request.method == 'POST':
        form = BankSoalForm(request.POST, instance=bank, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Bank Soal berhasil diperbarui.')
            return redirect('bank_soal_list')
    else:
        form = BankSoalForm(instance=bank, user=request.user)
    return render(request, 'exams/bank_soal_form.html', {'form': form, 'title': 'Edit Bank Soal'})

@login_required
def bank_soal_delete(request, pk):
    bank = get_object_or_404(BankSoal, pk=pk)
    if request.user.role == 'guru' and bank.guru != request.user:
        return redirect('login')
    
    if request.method == 'POST':
        bank.delete()
        messages.success(request, 'Bank Soal berhasil dihapus.')
        return redirect('bank_soal_list')
    return render(request, 'exams/bank_soal_confirm_delete.html', {'bank': bank})

@login_required
def bank_soal_detail(request, pk):
    bank = get_object_or_404(BankSoal, pk=pk)
    questions = bank.questions.all()
    return render(request, 'exams/bank_soal_detail.html', {'bank': bank, 'questions': questions})

@login_required
def butir_soal_create(request, pk):
    bank = get_object_or_404(BankSoal, pk=pk)
    if request.user.role == 'guru' and bank.guru != request.user:
        return redirect('login')

    if request.method == 'POST':
        form = ButirSoalForm(request.POST, request.FILES)
        if form.is_valid():
            soal = form.save(commit=False)
            soal.bank_soal = bank
            soal.save()
            messages.success(request, 'Butir soal berhasil ditambahkan.')
            return redirect('bank_soal_detail', pk=bank.pk)
    else:
        form = ButirSoalForm()
    return render(request, 'exams/butir_soal_form.html', {'form': form, 'title': 'Tambah Butir Soal', 'bank': bank})

@login_required
def butir_soal_edit(request, pk):
    soal = get_object_or_404(ButirSoal, pk=pk)
    bank = soal.bank_soal
    if request.user.role == 'guru' and bank.guru != request.user:
        return redirect('login')
        
    if request.method == 'POST':
        form = ButirSoalForm(request.POST, request.FILES, instance=soal)
        if form.is_valid():
            form.save()
            messages.success(request, 'Butir soal berhasil diperbarui.')
            return redirect('bank_soal_detail', pk=bank.pk)
    else:
        form = ButirSoalForm(instance=soal)
    return render(request, 'exams/butir_soal_form.html', {'form': form, 'title': 'Edit Butir Soal', 'bank': bank})

@login_required
def butir_soal_delete(request, pk):
    soal = get_object_or_404(ButirSoal, pk=pk)
    bank = soal.bank_soal
    if request.user.role == 'guru' and bank.guru != request.user:
        return redirect('login')
    
    if request.method == 'POST':
        soal.delete()
        messages.success(request, 'Butir soal berhasil dihapus.')
        return redirect('bank_soal_detail', pk=bank.pk)
    return render(request, 'exams/butir_soal_confirm_delete.html', {'soal': soal})

from docx import Document
from pypdf import PdfReader
import re

from .services import QuestionImportService

@login_required
def import_soal(request, pk):
    bank = get_object_or_404(BankSoal, pk=pk)
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        
        if uploaded_file.name.endswith('.xlsx'):
            service = QuestionImportService(uploaded_file)
            valid_data, errors = service.parse()
            
            # Store data in session for confirmation
            # Note: Serializing complex objects via session pickle might work, or convert to pure dicts
            request.session['import_data'] = valid_data
            
            return render(request, 'exams/import_preview.html', {
                'bank': bank,
                'valid_data': valid_data,
                'errors': errors
            })
            
        else:
            messages.error(request, 'Saat ini hanya mendukung format .xlsx untuk Smart Import.')
            return redirect('import_soal', pk=bank.pk)
            
    return render(request, 'exams/import_soal.html', {'bank': bank})

@login_required
def import_commit(request, pk):
    bank = get_object_or_404(BankSoal, pk=pk)
    if request.method == 'POST':
        data = request.session.get('import_data')
        if data:
            service = QuestionImportService(None) # Dummy init for save method, or better make save static
            service.save_to_bank(bank, data)
            
            del request.session['import_data']
            messages.success(request, f'Berhasil mengimpor {len(data)} soal.')
            return redirect('bank_soal_detail', pk=bank.pk)
        else:
            messages.error(request, 'Data sesi kadaluarsa. Silakan upload ulang.')
            return redirect('import_soal', pk=bank.pk)
    return redirect('bank_soal_detail', pk=bank.pk)

@login_required
def download_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Soal"
    
    # Headers matching the parser expectation
    headers = [
        "Pertanyaan", "Opsi A", "Opsi B", "Opsi C", "Opsi D", "Opsi E", 
        "Kunci (A-E)", "Bobot", "Tipe (PG/ESSAY)"
    ]
    ws.append(headers)
    
    # Sample PG
    ws.append([
        "Apa ibukota Indonesia?", "Jakarta", "Bandung", "Surabaya", "Medan", "Bali", 
        "A", 1, "PG"
    ])
    
    # Sample Essay
    ws.append([
        "Jelaskan pengertian fotosintesis!", "", "", "", "", "", 
        "", 5, "ESSAY"
    ])
    
    # Adjust column widths
    for col in ['A']:
        ws.column_dimensions[col].width = 40
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Template_Soal_CBT.xlsx'
    wb.save(response)
    return response

@login_required
def ujian_list(request):
    if request.user.role == 'guru':
        ujian_list = Ujian.objects.filter(bank_soal__guru=request.user)
    elif request.user.role == 'siswa':
        # Siswa hanya melihat ujian yang aktif dan sesuai kelasnya
        ujian_list = Ujian.objects.filter(kelas=request.user.kelas, aktif=True)
    else:
        ujian_list = Ujian.objects.all()

    # Filtering
    kelas_id = request.GET.get('kelas')
    waktu = request.GET.get('waktu')
    
    if kelas_id:
        ujian_list = ujian_list.filter(kelas__id=kelas_id)
        
    from django.utils import timezone
    now = timezone.now()
    
    if waktu == 'today':
        ujian_list = ujian_list.filter(waktu_mulai__date=now.date())
    elif waktu == 'upcoming':
        ujian_list = ujian_list.filter(waktu_mulai__gt=now)
    elif waktu == 'past':
        ujian_list = ujian_list.filter(waktu_mulai__lt=now)

    ujian_list = ujian_list.distinct().order_by('-waktu_mulai')
    
    # Get available classes for filter (only for guru/admin)
    available_classes = []
    if request.user.role in ['guru', 'admin']:
        from apps.academic.models import Kelas
        available_classes = Kelas.objects.all()

    return render(request, 'exams/ujian_list.html', {
        'ujian_list': ujian_list,
        'available_classes': available_classes
    })

@login_required
def ujian_create(request):
    if request.user.role != 'guru':
        return redirect('login')
        
    if request.method == 'POST':
        form = UjianForm(request.POST)
        if form.is_valid():
            ujian = form.save()
            messages.success(request, 'Jadwal ujian berhasil dibuat.')
            return redirect('ujian_list')
    else:
        form = UjianForm()
        # Filter bank soal by guru
        form.fields['bank_soal'].queryset = BankSoal.objects.filter(guru=request.user)
        
    return render(request, 'exams/ujian_form.html', {'form': form, 'title': 'Jadwalkan Ujian'})

@login_required
def ujian_edit(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    if request.method == 'POST':
        form = UjianForm(request.POST, instance=ujian)
        if form.is_valid():
            form.save()
            messages.success(request, 'Jadwal ujian berhasil diperbarui.')
            return redirect('ujian_list')
    else:
        form = UjianForm(instance=ujian)
        form.fields['bank_soal'].queryset = BankSoal.objects.filter(guru=request.user)
    return render(request, 'exams/ujian_form.html', {'form': form, 'title': 'Edit Jadwal Ujian'})

@login_required
def ujian_delete(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    if request.method == 'POST':
        ujian.delete()
        messages.success(request, 'Jadwal ujian berhasil dihapus.')
        return redirect('ujian_list')
    return render(request, 'exams/ujian_confirm_delete.html', {'ujian': ujian})

@login_required
def hasil_ujian(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    sesi_list = SesiUjian.objects.filter(ujian=ujian).order_by('-nilai')
    return render(request, 'exams/hasil_ujian.html', {'ujian': ujian, 'sesi_list': sesi_list})

@login_required
def konfirmasi_ujian(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    if request.user.role != 'siswa':
        return redirect('login')
    
    # Check existing session
    sesi = SesiUjian.objects.filter(ujian=ujian, siswa=request.user).first()
    if sesi:
        if sesi.is_finished:
            messages.info(request, 'Anda sudah menyelesaikan ujian ini.')
            return redirect('siswa_dashboard')
        return redirect('mulai_ujian', pk=pk)
    
    if request.method == 'POST':
        token_input = request.POST.get('token', '').upper()
        if token_input == ujian.token:
            # Create session
            SesiUjian.objects.create(
                ujian=ujian, 
                siswa=request.user,
                sisa_waktu=ujian.durasi * 60,
                score=0
            )
            return redirect('mulai_ujian', pk=pk)
        else:
            messages.error(request, 'Token salah.')
            
    return render(request, 'exams/konfirmasi_ujian.html', {'ujian': ujian})

@login_required
def mulai_ujian(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    sesi = get_object_or_404(SesiUjian, ujian=ujian, siswa=request.user)
    
    if sesi.is_finished:
        return redirect('siswa_dashboard')
        
    # Calculate remaining time
    now = timezone.now()
    elapsed = (now - sesi.waktu_mulai).total_seconds()
    duration_seconds = ujian.durasi * 60
    remaining_seconds = max(0, duration_seconds - elapsed)
    
    # Auto-finish if time is up
    if remaining_seconds <= 0:
        return redirect('selesai_ujian', pk=pk)
        
    # Get questions ordered by ID for consistent navigation
    questions = ujian.bank_soal.questions.all().order_by('id') 
    
    # Fetch existing answers
    saved_answers = {j.soal_id: j for j in JawabanSiswa.objects.filter(sesi=sesi)}
    
    # Prepare JSON data for JavaScript
    saved_answers_json = {}
    for soal_id, jawaban in saved_answers.items():
        saved_answers_json[str(soal_id)] = {
            'jawaban': jawaban.jawaban or '',
            'jawaban_essay': jawaban.jawaban_essay or '',
            'ragu_ragu': jawaban.ragu_ragu
        }
    
    questions_list = list(questions.values_list('id', flat=True))
    
    return render(request, 'exams/lembar_ujian.html', {
        'ujian': ujian,
        'sesi': sesi,
        'questions': questions,
        'saved_answers': saved_answers,
        'saved_answers_json': json.dumps(saved_answers_json),
        'questions_list': json.dumps(questions_list),
        'remaining_seconds': int(remaining_seconds)
    })

@login_required
def simpan_jawaban(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            sesi_id = data.get('sesi_id')
            soal_id = data.get('soal_id')
            jawaban = data.get('jawaban')
            ragu = data.get('ragu_ragu', False)
            
            sesi = get_object_or_404(SesiUjian, pk=sesi_id, siswa=request.user)
            if sesi.is_finished:
                return JsonResponse({'status': 'error', 'message': 'Ujian sudah selesai'})
                
            soal = get_object_or_404(ButirSoal, pk=soal_id)
            
            jawaban_obj, created = JawabanSiswa.objects.get_or_create(sesi=sesi, soal=soal)
            if 'jawaban_essay' in data:
                jawaban_obj.jawaban_essay = data['jawaban_essay']
        
            if 'jawaban' in data:
                jawaban_obj.jawaban = data['jawaban']
            
            jawaban_obj.ragu_ragu = data.get('ragu_ragu', False)
            jawaban_obj.save()
            
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid method'})

@login_required
def selesai_ujian(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    sesi = get_object_or_404(SesiUjian, ujian=ujian, siswa=request.user)
    
    if not sesi.is_finished:
        # Calculate Score PG
        total_score_pg = 0
        total_bobot = 0
        has_essay = False
        
        # Get all questions
        questions = ujian.bank_soal.questions.all()
        answers = JawabanSiswa.objects.filter(sesi=sesi)
        answers_map = {a.soal_id: a for a in answers}
        
        for q in questions:
            total_bobot += q.bobot
            if q.jenis_soal == 'ESSAY':
                has_essay = True
            else:
                # Logic PG
                student_ans_obj = answers_map.get(q.id)
                if student_ans_obj and student_ans_obj.jawaban == q.kunci_jawaban:
                    total_score_pg += q.bobot
                    # Update score on item level too
                    student_ans_obj.score = q.bobot
                    student_ans_obj.save()
                elif student_ans_obj:
                    student_ans_obj.score = 0
                    student_ans_obj.save()
                
        # Finalize Status
        sesi.is_finished = True
        sesi.waktu_selesai = timezone.now()
        
        if has_essay:
            sesi.status = 'WAITING_GRADE'
            # Temporarily store PG score, but final nilai waits for manual grading
            sesi.nilai = total_score_pg # Raw PG score
        else:
            sesi.status = 'GRADED'
            final_grade = 0
            if total_bobot > 0:
                final_grade = (total_score_pg / total_bobot) * 100
            sesi.nilai = round(final_grade, 2)
            
        sesi.save()
        
    messages.success(request, 'Ujian selesai.')
    return redirect('siswa_dashboard')

@login_required
def koreksi_list(request):
    if request.user.role != 'guru':
        return redirect('login')
        
    sesi_list = SesiUjian.objects.filter(
        ujian__bank_soal__guru=request.user,
        status='WAITING_GRADE',
        is_finished=True
    ).select_related('ujian', 'siswa').order_by('waktu_selesai')
    
    return render(request, 'exams/koreksi_list.html', {'sesi_list': sesi_list})

@login_required
def koreksi_detail(request, pk):
    sesi = get_object_or_404(SesiUjian, pk=pk)
    if request.user.role != 'guru' or sesi.ujian.bank_soal.guru != request.user:
        return redirect('login')
        
    if request.method == 'POST':
        total_score_essay = 0
        for key, value in request.POST.items():
            if key.startswith('score_'):
                jawaban_id = key.split('_')[1]
                try:
                    score = float(value)
                    jawaban = JawabanSiswa.objects.get(id=jawaban_id, sesi=sesi)
                    jawaban.score = score
                    jawaban.save()
                    total_score_essay += score
                except ValueError:
                    pass
        
        all_answers = sesi.jawaban_siswa.all()
        total_obtained = sum(a.score for a in all_answers)
        
        questions = sesi.ujian.bank_soal.questions.all()
        total_bobot = questions.aggregate(Sum('bobot'))['bobot__sum'] or 1
        
        final_grade = (total_obtained / total_bobot) * 100
        
        sesi.nilai = round(final_grade, 2)
        sesi.status = 'GRADED'
        sesi.save()
        
        messages.success(request, f'Nilai berhasil disimpan. Nilai Akhir: {sesi.nilai}')
        return redirect('koreksi_list')
        
    essay_answers = JawabanSiswa.objects.filter(
        sesi=sesi, 
        soal__jenis_soal='ESSAY'
    ).select_related('soal')
    
    return render(request, 'exams/koreksi_detail.html', {
        'sesi': sesi, 
        'essay_answers': essay_answers
    })

@login_required
def analisis_ujian(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    if request.user.role != 'guru' or ujian.bank_soal.guru != request.user:
        return redirect('login')

    questions = ujian.bank_soal.questions.all()
    sessions = SesiUjian.objects.filter(ujian=ujian, is_finished=True)
    student_count = sessions.count()
    
    analysis_data = []
    
    if student_count > 0:
        for q in questions:
            answers = JawabanSiswa.objects.filter(sesi__in=sessions, soal=q)
            total_score = answers.aggregate(Sum('score'))['score__sum'] or 0
            max_possible = q.bobot * student_count
            
            difficulty = 0
            if max_possible > 0:
                difficulty = total_score / max_possible
                
            if difficulty > 0.7:
                kategori = 'Mudah'
                badge = 'success'
            elif difficulty > 0.3:
                kategori = 'Sedang'
                badge = 'warning'
            else:
                kategori = 'Sukar'
                badge = 'danger'
                
            analysis_data.append({
                'soal': q,
                'total_score': total_score,
                'difficulty': round(difficulty, 2),
                'kategori': kategori,
                'badge': badge
            })
            
    return render(request, 'exams/analisis_ujian.html', {
        'ujian': ujian, 
        'analysis_data': analysis_data,
        'student_count': student_count
    })

@login_required
def analisis_export(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    if request.user.role != 'guru' or ujian.bank_soal.guru != request.user:
        return redirect('login')
        
    questions = ujian.bank_soal.questions.all()
    sessions = SesiUjian.objects.filter(ujian=ujian, is_finished=True)
    student_count = sessions.count()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analisis Butir Soal"
    
    headers = ['No', 'Pertanyaan', 'Tipe', 'Bobot', 'Total Skor Siswa', 'Indeks Kesukaran', 'Kategori']
    ws.append(headers)
    
    for idx, q in enumerate(questions, 1):
        answers = JawabanSiswa.objects.filter(sesi__in=sessions, soal=q)
        total_score = answers.aggregate(Sum('score'))['score__sum'] or 0
        max_possible = q.bobot * student_count if student_count > 0 else 1
        
        difficulty = 0
        if student_count > 0:
            difficulty = total_score / max_possible
            
        if difficulty > 0.7: kategori = 'Mudah'
        elif difficulty > 0.3: kategori = 'Sedang'
        else: kategori = 'Sukar'
        
        row = [
            idx,
            str(q.pertanyaan),
            q.jenis_soal,
            q.bobot,
            total_score,
            round(difficulty, 2),
            kategori
        ]
        ws.append(row)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Analisis_{ujian.nama_ujian}.xlsx'
    wb.save(response)
    return response

@login_required
def export_nilai_siswa(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    if request.user.role != 'guru' or ujian.bank_soal.guru != request.user:
        return redirect('login')
        
    sessions = SesiUjian.objects.filter(ujian=ujian, is_finished=True).select_related('siswa', 'siswa__kelas')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daftar Nilai"
    
    headers = ['No', 'NIS/Username', 'Nama Siswa', 'Kelas', 'Waktu Selesai', 'Status', 'Nilai Akhir']
    ws.append(headers)
    
    for idx, sesi in enumerate(sessions, 1):
        status_text = 'Sudah Dinilai' if sesi.status == 'GRADED' else 'Menunggu Koreksi'
        
        row = [
            idx,
            sesi.siswa.username,
            sesi.siswa.nama or sesi.siswa.username.replace('.', ' ').title(),
            sesi.siswa.kelas.nama if sesi.siswa.kelas else '-',
            sesi.waktu_selesai.strftime('%Y-%m-%d %H:%M') if sesi.waktu_selesai else '-',
            status_text,
            sesi.nilai
        ]
        ws.append(row)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Nilai_{ujian.nama_ujian}.xlsx'
    wb.save(response)
    return response


@login_required
def export_bank_soal(request, pk):
    bank = get_object_or_404(BankSoal, pk=pk)
    if request.user.role != 'guru' or bank.guru != request.user:
        return redirect('bank_soal_list')
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Soal"
    
    # Headers
    headers = [
        'Pertanyaan', 'Jenis Soal', 'Opsi A', 'Opsi B', 'Opsi C', 'Opsi D', 'Opsi E', 'Kunci', 'Bobot'
    ]
    ws.append(headers)
    
    for q in bank.questions.all():
        row = [
            q.pertanyaan,
            q.jenis_soal,
            q.opsi_a,
            q.opsi_b,
            q.opsi_c,
            q.opsi_d,
            q.opsi_e,
            q.kunci_jawaban,
            q.bobot
        ]
        ws.append(row)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Export_{bank.judul}.xlsx'
    wb.save(response)
    return response

@login_required
def regenerate_token(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    # Allow Guru or Admin
    if request.user.role == 'siswa':
        return redirect('login')
        
    import random, string
    new_token = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    ujian.token = new_token
    ujian.save()
    messages.success(request, f"Token ujian '{ujian.nama_ujian}' berhasil diperbarui: {new_token}")
    return redirect('ujian_list')

@login_required
def force_finish_exam(request, pk):
    sesi = get_object_or_404(SesiUjian, pk=pk)
    # Check permissions (teacher owner or admin)
    if request.user.role == 'siswa':
        return redirect('login')
        
    # Logic similar to selesai_ujian but forced
    sesi.waktu_selesai = timezone.now()
    sesi.is_finished = True
    
    # Calculate score logic (PG only usually for force finish)
    ujian = sesi.ujian
    essay_exists = ujian.bank_soal.questions.filter(jenis_soal='ESSAY').exists()
    
    # PG Grading
    questions = ujian.bank_soal.questions.filter(jenis_soal='PG')
    total_score = 0
    for q in questions:
        jawaban_siswa = JawabanSiswa.objects.filter(sesi=sesi, soal=q).first()
        if jawaban_siswa and jawaban_siswa.jawaban == q.kunci_jawaban:
            jawaban_siswa.score = q.bobot
            jawaban_siswa.save()
            total_score += q.bobot
        else:
            if jawaban_siswa:
                 # Ensure score is 0 if wrong
                 jawaban_siswa.score = 0
                 jawaban_siswa.save()

    if essay_exists:
        sesi.status = 'WAITING_GRADE'
        # temporary score only PG
        sesi.nilai = total_score
    else:
        sesi.status = 'GRADED'
        sesi.nilai = total_score
        
    sesi.save()
    messages.warning(request, f"Ujian siswa {sesi.siswa.nama} dipaksa selesai.")
    return redirect('ujian_list') # Ideally redirect to a monitoring page

@login_required
def reset_login(request, pk):
    sesi = get_object_or_404(SesiUjian, pk=pk)
    if request.user.role == 'siswa':
        return redirect('login')
        
    sesi.device_id = None # Clear device lock
    sesi.save()
    messages.success(request, f"Login reset untuk siswa {sesi.siswa.nama}.")
    return redirect('monitoring_ujian', pk=sesi.ujian.pk)

@login_required
def chart_analisis(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    questions = ujian.bank_soal.questions.all()
    
    mudah = 0
    sedang = 0
    sukar = 0
    
    for q in questions:
        benar = JawabanSiswa.objects.filter(soal=q, jawaban=q.kunci_jawaban).count()
        total = JawabanSiswa.objects.filter(soal=q).count()
        
        if total > 0:
            ratio = benar / total
            if ratio > 0.7:
                mudah += 1
            elif ratio > 0.3:
                sedang += 1
            else:
                sukar += 1
        else:
            sedang += 1 # Default to medium if no data
            
    return JsonResponse({
        'labels': ['Mudah', 'Sedang', 'Sukar'],
        'data': [mudah, sedang, sukar],
        'colors': ['#2fb344', '#f59f00', '#d63939']
    })

@login_required
def cetak_kartu_ujian(request, pk):
    sesi = get_object_or_404(SesiUjian, pk=pk)
    # Security: Only teacher of this exam or the student themselves can print
    if request.user.role == 'siswa' and sesi.siswa != request.user:
        return redirect('login')
    if request.user.role == 'guru' and sesi.ujian.bank_soal.guru != request.user:
        return redirect('login')
        
    # Calculate stats
    total_dijawab = JawabanSiswa.objects.filter(sesi=sesi).count()
    total_benar = 0
    total_salah = 0
    
    # Simple calculation for PG
    jawaban_list = JawabanSiswa.objects.filter(sesi=sesi)
    for j in jawaban_list:
        if j.soal.jenis_soal == 'PG':
            if j.jawaban == j.soal.kunci_jawaban:
                total_benar += 1
            else:
                total_salah += 1
                
    # Format student name
    nama_siswa = sesi.siswa.nama
    if not nama_siswa:
        nama_siswa = sesi.siswa.username.replace('.', ' ').title()
        
    nomor_induk = sesi.siswa.nisn or sesi.siswa.username
        
    return render(request, 'exams/kartu_ujian.html', {
        'sesi': sesi,
        'nama_siswa': nama_siswa,
        'nomor_induk': nomor_induk,
        'total_dijawab': total_dijawab,
        'total_benar': total_benar,
        'total_salah': total_salah
    })

@login_required
def monitoring_ujian(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    if request.user.role not in ['guru', 'admin', 'proktor']:
         return redirect('login')
         
    # Filter sessions based on role
    active_sessions = SesiUjian.objects.filter(ujian=ujian, is_finished=False)
    
    is_owner = (request.user == ujian.bank_soal.guru)
    
    # If proktor OR (teacher and NOT owner)
    if request.user.role == 'proktor' or (request.user.role == 'guru' and not is_owner):
        from .models import JadwalPengawas
        assigned_classes = JadwalPengawas.objects.filter(
            ujian=ujian, 
            proktor=request.user
        ).values_list('kelas', flat=True)
        
        # Security: If no assignment found
        if not assigned_classes:
            messages.error(request, 'Anda tidak ditugaskan di ujian ini.')
            return redirect('guru_dashboard' if request.user.role == 'guru' else 'proktor_dashboard')
            
        active_sessions = active_sessions.filter(siswa__kelas__in=assigned_classes)
    
    # AJAX Handler
    if request.GET.get('ajax') == '1':
        data = {
            'sessions': []
        }
        for s in active_sessions:
            # Calculate progress logic
            elapsed = (timezone.now() - s.waktu_mulai).total_seconds()
            duration_seconds = ujian.durasi * 60
            progress_percent = 0
            if duration_seconds > 0:
                progress_percent = min(100, (elapsed / duration_seconds) * 100)
            
            # Recalculate sisa_waktu just in case
            remaining = max(0, int(duration_seconds - elapsed))
            hours, remainder = divmod(remaining, 3600)
            minutes, seconds = divmod(remainder, 60)
            time_str = f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
            
            data['sessions'].append({
                'id': s.id,
                'username': s.siswa.username,
                'nama': s.siswa.nama,
                'is_finished': s.is_finished,
                'device_id': s.device_id,
                'nilai': s.nilai,
                'sisa_waktu': remaining,
                'sisa_waktu_str': time_str,
                'progress_percent': int(progress_percent)
            })
        return JsonResponse(data)
    
    return render(request, 'exams/monitoring_ujian.html', {
        'ujian': ujian,
        'active_sessions': active_sessions
    })

@login_required
def kartu_peserta_ujian(request, pk):
    """Generate participant cards with login credentials for admin/proctor to print"""
    ujian = get_object_or_404(Ujian, pk=pk)
    
    # Security: Only admin, proktor, or teacher owner can access
    if request.user.role == 'siswa':
        return redirect('login')
    if request.user.role == 'guru' and ujian.bank_soal.guru != request.user:
        return redirect('login')
    # proktor and admin can access all exams
    
    # Get all students from exam classes
    from apps.accounts.models import CustomUser
    kelas_ids = ujian.kelas.values_list('id', flat=True)
    peserta_list = CustomUser.objects.filter(
        role='siswa',
        kelas_id__in=kelas_ids
    ).select_related('kelas').order_by('kelas__nama', 'username')
    
    return render(request, 'exams/kartu_peserta_ujian.html', {
        'ujian': ujian,
        'peserta_list': peserta_list
    })

@login_required
def atur_pengawas(request, pk):
    ujian = get_object_or_404(Ujian, pk=pk)
    # Security: Only Admin can assign proctors
    if request.user.role != 'admin':
        messages.error(request, 'Hanya Admin yang dapat mengatur pengawas.')
        return redirect('ujian_list')
    
    from apps.exams.models import JadwalPengawas
    from apps.accounts.models import CustomUser
    
    classes = ujian.kelas.all()
    teachers = CustomUser.objects.filter(role='guru').order_by('nama')
    
    if request.method == 'POST':
        for kelas_obj in classes:
            # We use 'proctor_' prefix in form to match existing/generic name, or can change to 'teacher_'
            # Let's keep input name 'proctor_' to avoid changing too many things if not needed, 
            # OR better, change to 'supervisor_' since it's generic.
            # But the HTML has 'proctor_' in previous steps. I will check logic below.
            # The model field is still 'proktor'.
            supervisor_id = request.POST.get(f'supervisor_{kelas_obj.id}')
            if supervisor_id:
                # Update or Create assignment
                JadwalPengawas.objects.update_or_create(
                    ujian=ujian,
                    kelas=kelas_obj,
                    defaults={'proktor_id': supervisor_id}
                )
            else:
                # If selection is cleared (empty), remove assignment
                JadwalPengawas.objects.filter(ujian=ujian, kelas=kelas_obj).delete()
                
        messages.success(request, 'Jadwal pengawas (Guru) berhasil disimpan.')
        return redirect('atur_pengawas', pk=pk)

    # Dictionary of existing assignments: {kelas_id: proctor_id}
    existing_assignments = {
        jp.kelas_id: jp.proktor_id 
        for jp in JadwalPengawas.objects.filter(ujian=ujian)
    }
    
    # Prepare class_data with pre-computed selection
    class_data = []
    for kelas in classes:
        assigned_id = existing_assignments.get(kelas.id)
        teacher_options = []
        for t in teachers:
            teacher_options.append({
                'id': t.id,
                'nama': t.nama,
                'username': t.username,
                'is_selected': (t.id == assigned_id)
            })
        class_data.append({
            'kelas': kelas,
            'teachers': teacher_options
        })
    
    return render(request, 'exams/atur_pengawas.html', {
        'ujian': ujian,
        'class_data': class_data,
    })

@login_required
def jadwal_pengawas_list(request):
    if request.user.role != 'admin':
        messages.error(request, 'Akses ditolak.')
        return redirect('login')
    
    jadwal_pengawas = JadwalPengawas.objects.select_related('ujian', 'kelas', 'proktor').all().order_by('-id')
    return render(request, 'exams/jadwal_pengawas_list.html', {'jadwal_pengawas': jadwal_pengawas})

@login_required
def jadwal_pengawas_create(request):
    if request.user.role != 'admin':
        messages.error(request, 'Akses ditolak.')
        return redirect('login')

    if request.method == 'POST':
        form = JadwalPengawasForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Jadwal pengawas berhasil ditambahkan.')
            return redirect('jadwal_pengawas_list')
    else:
        form = JadwalPengawasForm()
    
    return render(request, 'exams/jadwal_pengawas_form.html', {'form': form})

@login_required
def jadwal_pengawas_edit(request, pk):
    if request.user.role != 'admin':
        messages.error(request, 'Akses ditolak.')
        return redirect('login')
        
    jadwal = get_object_or_404(JadwalPengawas, pk=pk)
    
    if request.method == 'POST':
        form = JadwalPengawasForm(request.POST, instance=jadwal)
        if form.is_valid():
            form.save()
            messages.success(request, 'Jadwal pengawas berhasil diperbarui.')
            return redirect('jadwal_pengawas_list')
    else:
        form = JadwalPengawasForm(instance=jadwal)
    
    return render(request, 'exams/jadwal_pengawas_form.html', {'form': form})

@login_required
def jadwal_pengawas_delete(request, pk):
    if request.user.role != 'admin':
        messages.error(request, 'Akses ditolak.')
        return redirect('login')
        
    jadwal = get_object_or_404(JadwalPengawas, pk=pk)
    
    if request.method == 'POST':
        jadwal.delete()
        messages.success(request, 'Jadwal pengawas berhasil dihapus.')
        return redirect('jadwal_pengawas_list')
    
    return render(request, 'exams/jadwal_pengawas_confirm_delete.html', {'object': jadwal})

# ============================================
# PDF EXPORT FUNCTIONS
# ============================================

@login_required
def analisis_export_pdf(request, pk):
    """Export analisis butir soal to PDF with logo and professional design"""
    from .pdf_generator import PDFGenerator
    
    ujian = get_object_or_404(Ujian, pk=pk)
    if request.user.role != 'guru' or ujian.bank_soal.guru != request.user:
        return redirect('login')
    
    questions = ujian.bank_soal.questions.all()
    sessions = SesiUjian.objects.filter(ujian=ujian, is_finished=True)
    student_count = sessions.count()
    
    # Create PDF buffer
    buffer = BytesIO()
    pdf = PDFGenerator(buffer, title=f"Analisis Butir Soal - {ujian.nama_ujian}")
    
    # Add header
    pdf.add_header(
        school_name="SMAN 1 KRAGILAN",
        school_address="Banten",
        report_title="ANALISIS BUTIR SOAL"
    )
    
    # Add info section
    info_dict = {
        "Mata Pelajaran": ujian.bank_soal.mapel.nama,
        "Nama Ujian": ujian.nama_ujian,
        "Tanggal Ujian": ujian.waktu_mulai.strftime('%d %B %Y, %H:%M'),
        "Jumlah Soal": questions.count(),
        "Jumlah Peserta": student_count
    }
    pdf.add_info_section(info_dict)
    
    # Prepare table data
    table_data = [
        ['No', 'Pertanyaan', 'Tipe', 'Bobot', 'Total Skor', 'Indeks (P)', 'Kategori']
    ]
    
    mudah_count = 0
    sedang_count = 0
    sukar_count = 0
    
    for idx, q in enumerate(questions, 1):
        answers = JawabanSiswa.objects.filter(sesi__in=sessions, soal=q)
        total_score = answers.aggregate(Sum('score'))['score__sum'] or 0
        max_possible = q.bobot * student_count if student_count > 0 else 1
        
        difficulty = 0
        if student_count > 0:
            difficulty = total_score / max_possible
        
        if difficulty > 0.7:
            kategori = 'Mudah'
            mudah_count += 1
        elif difficulty > 0.3:
            kategori = 'Sedang'
            sedang_count += 1
        else:
            kategori = 'Sukar'
            sukar_count += 1
        
        # Truncate question text if too long
        pertanyaan_text = str(q.pertanyaan)
        if len(pertanyaan_text) > 50:
            pertanyaan_text = pertanyaan_text[:47] + "..."
        
        row = [
            str(idx),
            pertanyaan_text,
            q.jenis_soal,
            str(q.bobot),
            str(total_score),
            f"{difficulty:.2f}",
            kategori
        ]
        table_data.append(row)
    
    # Add table with custom column widths
    col_widths = [1*cm, 6*cm, 2*cm, 1.5*cm, 2*cm, 2*cm, 2*cm]
    pdf.add_table(table_data, col_widths=col_widths)
    
    # Add statistics
    stats_dict = {
        "Soal Mudah (P > 0.70)": f"{mudah_count} soal",
        "Soal Sedang (0.31 - 0.70)": f"{sedang_count} soal",
        "Soal Sukar (P ≤ 0.30)": f"{sukar_count} soal"
    }
    pdf.add_statistics(stats_dict)
    
    # Add footer
    pdf.add_footer()
    
    # Build PDF
    pdf.build()
    buffer.seek(0)
    
    # Return as response
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Analisis_{ujian.nama_ujian}.pdf'
    return response


@login_required
def export_nilai_siswa_pdf(request, pk):
    """Export daftar nilai siswa to PDF with logo and statistics"""
    from .pdf_generator import PDFGenerator
    
    ujian = get_object_or_404(Ujian, pk=pk)
    if request.user.role != 'guru' or ujian.bank_soal.guru != request.user:
        return redirect('login')
    
    sessions = SesiUjian.objects.filter(ujian=ujian, is_finished=True).select_related('siswa', 'siswa__kelas')
    
    # Create PDF buffer
    buffer = BytesIO()
    pdf = PDFGenerator(buffer, title=f"Daftar Nilai - {ujian.nama_ujian}")
    
    # Add header
    pdf.add_header(
        school_name="SMAN 1 KRAGILAN",
        school_address="Banten",
        report_title="DAFTAR NILAI UJIAN"
    )
    
    # Add info section
    info_dict = {
        "Mata Pelajaran": ujian.bank_soal.mapel.nama,
        "Nama Ujian": ujian.nama_ujian,
        "Tanggal Ujian": ujian.waktu_mulai.strftime('%d %B %Y, %H:%M'),
        "Durasi": f"{ujian.durasi} menit",
        "Jumlah Peserta": sessions.count()
    }
    pdf.add_info_section(info_dict)
    
    # Prepare table data
    table_data = [
        ['No', 'NIS/Username', 'Nama Siswa', 'Kelas', 'Waktu Selesai', 'Status', 'Nilai']
    ]
    
    nilai_list = []
    for idx, sesi in enumerate(sessions, 1):
        status_text = 'Dinilai' if sesi.status == 'GRADED' else 'Menunggu'
        waktu_selesai = sesi.waktu_selesai.strftime('%d/%m/%Y %H:%M') if sesi.waktu_selesai else '-'
        
        # Format nama: replace dots with spaces and title case
        nama_raw = sesi.siswa.nama or sesi.siswa.username
        nama_formatted = nama_raw.replace('.', ' ').title()
        
        row = [
            str(idx),
            sesi.siswa.username,
            nama_formatted,
            sesi.siswa.kelas.nama if sesi.siswa.kelas else '-',
            waktu_selesai,
            status_text,
            f"{sesi.nilai:.2f}"
        ]
        table_data.append(row)
        
        if sesi.status == 'GRADED':
            nilai_list.append(sesi.nilai)
    
    # Add table
    col_widths = [1*cm, 2.5*cm, 4*cm, 2*cm, 3*cm, 2*cm, 2*cm]
    pdf.add_table(table_data, col_widths=col_widths)
    
    # Calculate and add statistics
    if nilai_list:
        rata_rata = sum(nilai_list) / len(nilai_list)
        tertinggi = max(nilai_list)
        terendah = min(nilai_list)
        
        # Count passing students (assuming KKM = 75)
        lulus = sum(1 for n in nilai_list if n >= 75)
        persen_lulus = (lulus / len(nilai_list)) * 100 if nilai_list else 0
        
        stats_dict = {
            "Rata-rata Kelas": f"{rata_rata:.2f}",
            "Nilai Tertinggi": f"{tertinggi:.2f}",
            "Nilai Terendah": f"{terendah:.2f}",
            "Jumlah Lulus (≥75)": f"{lulus} siswa ({persen_lulus:.1f}%)"
        }
        pdf.add_statistics(stats_dict)
    
    # Add footer
    pdf.add_footer()
    
    # Build PDF
    pdf.build()
    buffer.seek(0)
    
    # Return as response
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Nilai_{ujian.nama_ujian}.pdf'
    return response

